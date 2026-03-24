"""
Student Import Service for Grade Portal.

Handles bulk student imports from Excel/CSV files.
"""
import io
import re
from datetime import datetime, timezone

from app.extensions import db
from app.models.student import Student
from app.models.section import Section
from app.models.user import User


def import_students_from_excel(file_stream) -> dict:
    """
    Parse and import students from an Excel (.xlsx) file.
    Expected columns: student_id, full_name, year_level, section, gmail
    Optional columns: age, address, contact_number, gender

    Returns: {'imported': N, 'skipped': M, 'errors': [...]}
    All-or-nothing: if there are ANY errors, no students are written.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            'imported': 0,
            'skipped': 0,
            'errors': ['openpyxl library not installed. Please install it with: pip install openpyxl'],
        }

    try:
        # Load workbook from file stream
        wb = load_workbook(filename=io.BytesIO(file_stream.read()), read_only=True, data_only=True)
        ws = wb.active  # Use first/active sheet

        if ws is None:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': ['Excel file has no worksheets.'],
            }

        # Get all rows as list
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': ['Excel file must have a header row and at least one data row.'],
            }

        # Parse header row (first row)
        header = [str(cell).strip().lower() if cell else '' for cell in rows[0]]

        # Map column names (flexible matching)
        col_map = {}
        for i, col_name in enumerate(header):
            col_clean = col_name.replace('_', ' ').strip()
            if col_clean in ('student_id', 'studentid', 'student id', 'id'):
                col_map['student_id'] = i
            elif col_clean in ('full_name', 'fullname', 'full name', 'name'):
                col_map['full_name'] = i
            elif col_clean in ('year_level', 'yearlevel', 'year level', 'year', 'level'):
                col_map['year_level'] = i
            elif col_clean in ('section', 'section_name', 'section name'):
                col_map['section'] = i
            elif col_clean in ('gmail', 'email', 'e-mail'):
                col_map['gmail'] = i
            elif col_clean in ('age',):
                col_map['age'] = i
            elif col_clean in ('address',):
                col_map['address'] = i
            elif col_clean in ('contact_number', 'contactnumber', 'contact number', 'contact', 'phone'):
                col_map['contact_number'] = i
            elif col_clean in ('gender', 'sex'):
                col_map['gender'] = i

        required_cols = {'student_id', 'full_name', 'year_level'}
        missing_cols = required_cols - set(col_map.keys())
        if missing_cols:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': [
                    f'Excel file must have columns: student_id, full_name, year_level. '
                    f'Missing: {", ".join(sorted(missing_cols))}. '
                    f'Found headers: {", ".join(header)}'
                ],
            }

        wb.close()

    except Exception as e:
        return {
            'imported': 0,
            'skipped': 0,
            'errors': [f'Error reading Excel file: {str(e)}'],
        }

    # Process data rows
    data_rows = rows[1:]  # Skip header
    return _validate_and_save_students(data_rows, col_map)


def import_students_from_csv(file_stream) -> dict:
    """
    Parse and import students from a CSV file.
    Expected columns: student_id, full_name, year_level, section, gmail
    Optional columns: age, address, contact_number, gender

    Returns: {'imported': N, 'skipped': M, 'errors': [...]}
    All-or-nothing: if there are ANY errors, no students are written.
    """
    import csv

    try:
        # Read CSV from stream
        file_stream.seek(0)
        content = file_stream.read().decode('utf-8-sig')  # Support UTF-8 with BOM
        reader = csv.DictReader(io.StringIO(content))

        # Normalize headers
        fieldnames = [field.strip().lower() if field else '' for field in reader.fieldnames or []]

        # Map column names
        col_map = {}
        for i, col_name in enumerate(fieldnames):
            col_clean = col_name.replace('_', ' ').strip()
            if col_clean in ('student_id', 'studentid', 'student id', 'id'):
                col_map['student_id'] = i
            elif col_clean in ('full_name', 'fullname', 'full name', 'name'):
                col_map['full_name'] = i
            elif col_clean in ('year_level', 'yearlevel', 'year level', 'year', 'level'):
                col_map['year_level'] = i
            elif col_clean in ('section', 'section_name', 'section name'):
                col_map['section'] = i
            elif col_clean in ('gmail', 'email', 'e-mail'):
                col_map['gmail'] = i
            elif col_clean in ('age',):
                col_map['age'] = i
            elif col_clean in ('address',):
                col_map['address'] = i
            elif col_clean in ('contact_number', 'contactnumber', 'contact number', 'contact', 'phone'):
                col_map['contact_number'] = i
            elif col_clean in ('gender', 'sex'):
                col_map['gender'] = i

        required_cols = {'student_id', 'full_name', 'year_level'}
        missing_cols = required_cols - set(col_map.keys())
        if missing_cols:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': [
                    f'CSV file must have columns: student_id, full_name, year_level. '
                    f'Missing: {", ".join(sorted(missing_cols))}. '
                    f'Found headers: {", ".join(fieldnames)}'
                ],
            }

        # Collect all rows as lists
        file_stream.seek(0)
        content = file_stream.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)[1:]  # Skip header

        return _validate_and_save_students(rows, col_map)

    except Exception as e:
        return {
            'imported': 0,
            'skipped': 0,
            'errors': [f'Error reading CSV file: {str(e)}'],
        }


def _validate_and_save_students(data_rows: list, col_map: dict) -> dict:
    """
    Common validation and save logic for Excel/CSV imports.

    Args:
        data_rows: List of row tuples/lists
        col_map: Dict mapping column names to indices

    Returns: {'imported': N, 'skipped': M, 'errors': [...]}
    """
    errors = []
    pending = []  # list of (student_data)
    seen_student_ids = set()
    seen_gmails = set()

    for i, row in enumerate(data_rows, start=2):  # row 1 is header
        try:
            # Extract required values
            student_id = str(row[col_map['student_id']] or '').strip()
            full_name = str(row[col_map['full_name']] or '').strip()
            year_level_raw = str(row[col_map['year_level']] or '').strip()

            # Clean up values
            student_id = re.sub(r'\s+', '', student_id).upper()
            full_name = ' '.join(full_name.split())  # Normalize spaces

            if not student_id or not full_name or not year_level_raw:
                errors.append(f'Row {i}: missing required value(s) (student_id, full_name, or year_level).')
                continue

            # Validate year level
            try:
                year_level = int(year_level_raw)
                if year_level not in [1, 2, 3, 4]:
                    errors.append(f'Row {i}: year_level must be 1, 2, 3, or 4. Got: {year_level_raw}')
                    continue
            except ValueError:
                errors.append(f'Row {i}: year_level "{year_level_raw}" is not a valid number.')
                continue

            # Check for duplicates in file
            if student_id in seen_student_ids:
                errors.append(f'Row {i}: duplicate student_id "{student_id}" in file.')
                continue
            seen_student_ids.add(student_id)

            # Check if student_id already exists in database
            existing_student = Student.query.filter_by(student_id=student_id).first()
            if existing_student:
                errors.append(f'Row {i}: student_id "{student_id}" already exists in database.')
                continue

            # Extract optional values
            section_name = str(row[col_map['section']] or '').strip() if 'section' in col_map else None
            gmail = str(row[col_map['gmail']] or '').strip().lower() if 'gmail' in col_map else None
            age = None
            if 'age' in col_map and row[col_map['age']]:
                try:
                    age = int(row[col_map['age']])
                except (ValueError, TypeError):
                    pass  # Ignore invalid age values

            address = str(row[col_map['address']] or '').strip() if 'address' in col_map else None
            contact_number = str(row[col_map['contact_number']] or '').strip() if 'contact_number' in col_map else None
            gender = str(row[col_map['gender']] or '').strip() if 'gender' in col_map else None

            # Validate gmail uniqueness if provided
            if gmail:
                if gmail in seen_gmails:
                    errors.append(f'Row {i}: duplicate gmail "{gmail}" in file.')
                    continue
                seen_gmails.add(gmail)

                # Check if email already exists
                existing_user = User.query.filter_by(email=gmail).first()
                if existing_user:
                    errors.append(f'Row {i}: email "{gmail}" already exists in database.')
                    continue

            # Validate section if provided
            section_id = None
            if section_name:
                section = Section.query.filter_by(display_name=section_name).first()
                if not section:
                    errors.append(f'Row {i}: section "{section_name}" not found. Please create it first.')
                    continue
                section_id = section.id

            # Build student data
            student_data = {
                'student_id': student_id,
                'full_name': full_name,
                'year_level': year_level,
                'section_id': section_id,
                'gmail': gmail,
                'age': age,
                'address': address,
                'contact_number': contact_number,
                'gender': gender,
            }

            pending.append(student_data)

        except (IndexError, TypeError) as e:
            errors.append(f'Row {i}: error reading row data - {str(e)}')
            continue

    # All-or-nothing: if any errors, don't save anything
    if errors:
        return {'imported': 0, 'skipped': 0, 'errors': errors}

    # Create all students
    now = datetime.now(timezone.utc)
    for student_data in pending:
        student = Student(
            user_id=None,  # No linked user account yet
            **student_data,
            created_at=now,
            updated_at=now,
        )
        db.session.add(student)

    db.session.commit()
    return {'imported': len(pending), 'skipped': 0, 'errors': []}

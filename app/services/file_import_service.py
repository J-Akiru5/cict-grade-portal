"""
File Import Service for Grade Portal.

Handles PDF and Excel grade imports, reusing validation logic from CSV import.
"""
import io
import re
from datetime import datetime, timezone

from app.extensions import db
from app.models.student import Student
from app.models.subject import Subject
from app.models.enrollment import Enrollment
from app.models.grade import Grade


# Valid grade values and remarks (must match admin_service.py)
VALID_GRADES = {1.0, 1.25, 1.5, 1.75, 2.0, 2.25, 2.5, 2.75, 3.0, 5.0}
VALID_REMARKS = {'INC', 'DRP'}


def import_grades_from_excel(
    file_stream,
    semester: str,
    academic_year: str,
    actor_user,
) -> dict:
    """
    Parse and import grades from an Excel (.xlsx) file.
    Expected columns: student_id, subject_code, grade
    'grade' can be a float (1.0–5.0) or 'INC'/'DRP'.

    Returns: {'imported': N, 'skipped': M, 'errors': [...]}
    All-or-nothing: if there are ANY errors, no grades are written.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            'imported': 0,
            'skipped': 0,
            'errors': ['openpyxl library not installed. Please install it with: pip install openpyxl'],
        }

    try:
        # Load workbook from file stream
        wb = load_workbook(filename=io.BytesIO(file_stream.read()), read_only=True, data_only=True)
        ws = wb.active  # Use first/active sheet

        if ws is None:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': ['Excel file has no worksheets.'],
            }

        # Get all rows as list
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': ['Excel file must have a header row and at least one data row.'],
            }

        # Parse header row (first row)
        header = [str(cell).strip().lower() if cell else '' for cell in rows[0]]

        # Map column names (flexible matching)
        col_map = {}
        for i, col_name in enumerate(header):
            if col_name in ('student_id', 'studentid', 'student id', 'id'):
                col_map['student_id'] = i
            elif col_name in ('subject_code', 'subjectcode', 'subject code', 'code', 'subject'):
                col_map['subject_code'] = i
            elif col_name in ('grade', 'grades', 'final grade', 'final_grade'):
                col_map['grade'] = i

        required_cols = {'student_id', 'subject_code', 'grade'}
        missing_cols = required_cols - set(col_map.keys())
        if missing_cols:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': [
                    f'Excel file must have columns: student_id, subject_code, grade. '
                    f'Missing: {", ".join(sorted(missing_cols))}. '
                    f'Found headers: {", ".join(header)}'
                ],
            }

        wb.close()

    except Exception as e:
        return {
            'imported': 0,
            'skipped': 0,
            'errors': [f'Error reading Excel file: {str(e)}'],
        }

    # Process data rows
    data_rows = rows[1:]  # Skip header
    return _validate_and_save_grades(data_rows, col_map, semester, academic_year, actor_user, 'Excel')


def import_grades_from_pdf(
    file_stream,
    semester: str,
    academic_year: str,
    actor_user,
) -> dict:
    """
    Extract and import grades from a PDF file containing grade tables.
    Uses pdfplumber for table extraction.

    Returns: {'imported': N, 'skipped': M, 'errors': [...]}
    All-or-nothing: if there are ANY errors, no grades are written.
    """
    try:
        import pdfplumber
    except ImportError:
        return {
            'imported': 0,
            'skipped': 0,
            'errors': ['pdfplumber library not installed. Please install it with: pip install pdfplumber'],
        }

    try:
        # Read PDF from stream
        pdf_bytes = io.BytesIO(file_stream.read())
        all_tables = []

        with pdfplumber.open(pdf_bytes) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                # Extract tables from each page
                tables = page.extract_tables()
                for table in tables:
                    if table and len(table) > 1:  # Must have at least header + 1 row
                        all_tables.append((page_num, table))

        if not all_tables:
            # Try extracting text if no tables found
            return {
                'imported': 0,
                'skipped': 0,
                'errors': [
                    'No grade tables found in PDF. Please ensure the PDF contains '
                    'tabular data with columns: Student ID, Subject Code, Grade.'
                ],
            }

        # Process extracted tables
        extracted_rows = []
        col_map = None

        for page_num, table in all_tables:
            # Parse header from first row
            header = [str(cell).strip().lower() if cell else '' for cell in table[0]]

            # Try to map columns
            table_col_map = {}
            for i, col_name in enumerate(header):
                col_clean = col_name.replace('\n', ' ').strip()
                if any(term in col_clean for term in ('student id', 'student_id', 'studentid', 'id no')):
                    table_col_map['student_id'] = i
                elif any(term in col_clean for term in ('subject code', 'subject_code', 'subjectcode', 'course code', 'code')):
                    table_col_map['subject_code'] = i
                elif any(term in col_clean for term in ('grade', 'final grade', 'final_grade', 'rating')):
                    table_col_map['grade'] = i

            # Check if this table has the required columns
            if set(table_col_map.keys()) >= {'student_id', 'subject_code', 'grade'}:
                if col_map is None:
                    col_map = table_col_map

                # Add data rows (skip header)
                for row in table[1:]:
                    if row and any(cell for cell in row):  # Skip empty rows
                        extracted_rows.append(row)

        if not extracted_rows or col_map is None:
            return {
                'imported': 0,
                'skipped': 0,
                'errors': [
                    'Could not extract grade data from PDF tables. '
                    'Please ensure tables have columns: Student ID, Subject Code, Grade.'
                ],
            }

    except Exception as e:
        return {
            'imported': 0,
            'skipped': 0,
            'errors': [f'Error reading PDF file: {str(e)}'],
        }

    return _validate_and_save_grades(extracted_rows, col_map, semester, academic_year, actor_user, 'PDF')


def _validate_and_save_grades(
    data_rows: list,
    col_map: dict,
    semester: str,
    academic_year: str,
    actor_user,
    source_type: str,
) -> dict:
    """
    Common validation and save logic for Excel/PDF imports.
    Reuses the same validation rules as CSV import.

    Args:
        data_rows: List of row tuples/lists
        col_map: Dict mapping column names to indices
        semester: Target semester
        academic_year: Target academic year
        actor_user: User performing the import
        source_type: 'Excel' or 'PDF' for error messages

    Returns: {'imported': N, 'skipped': M, 'errors': [...]}
    """
    errors = []
    pending = []  # list of (enrollment, grade_value, remarks)

    for i, row in enumerate(data_rows, start=2):  # row 1 is header
        try:
            # Extract values using column mapping
            sid = str(row[col_map['student_id']] or '').strip()
            code = str(row[col_map['subject_code']] or '').strip().upper()
            raw_grade = str(row[col_map['grade']] or '').strip().upper()

            # Clean up values (remove newlines, extra spaces)
            sid = re.sub(r'\s+', '', sid)
            code = re.sub(r'\s+', ' ', code).strip()
            raw_grade = re.sub(r'\s+', '', raw_grade)

            if not sid or not code or not raw_grade:
                errors.append(f'{source_type} Row {i}: missing value(s).')
                continue

            # Validate student
            student = Student.query.filter_by(student_id=sid).first()
            if not student:
                errors.append(f'{source_type} Row {i}: student_id "{sid}" not found.')
                continue

            # Validate subject (try both subject_code and code fields)
            subject = Subject.query.filter(
                (Subject.subject_code == code) | (Subject.code == code)
            ).first()
            if not subject:
                errors.append(f'{source_type} Row {i}: subject_code "{code}" not found.')
                continue

            # Validate enrollment
            enrollment = Enrollment.query.filter_by(
                student_id=student.id,
                subject_id=subject.id,
                semester=semester,
                academic_year=academic_year,
            ).first()
            if not enrollment:
                errors.append(
                    f'{source_type} Row {i}: no enrollment for {sid} in {code} '
                    f'({semester} {academic_year}).'
                )
                continue

            # Parse grade value
            if raw_grade in VALID_REMARKS:
                grade_value, remarks = None, raw_grade
            else:
                try:
                    gv = round(float(raw_grade), 2)
                    if gv not in VALID_GRADES:
                        errors.append(
                            f'{source_type} Row {i}: "{raw_grade}" is not a valid grade value '
                            f'(allowed: {sorted(VALID_GRADES)} or INC/DRP).'
                        )
                        continue
                    grade_value, remarks = gv, None
                except ValueError:
                    errors.append(f'{source_type} Row {i}: "{raw_grade}" is not a valid grade.')
                    continue

            pending.append((enrollment, grade_value, remarks))

        except (IndexError, TypeError) as e:
            errors.append(f'{source_type} Row {i}: error reading row data - {str(e)}')
            continue

    # All-or-nothing: if any errors, don't save anything
    if errors:
        return {'imported': 0, 'skipped': 0, 'errors': errors}

    # Apply all grades
    now = datetime.now(timezone.utc)
    for enrollment, grade_value, remarks in pending:
        grade = enrollment.grade
        if grade is None:
            grade = Grade(enrollment_id=enrollment.id)
            db.session.add(grade)
        grade.grade_value = grade_value
        grade.remarks = remarks
        grade.date_encoded = now
        grade.encoded_by_id = actor_user.id

    db.session.commit()
    return {'imported': len(pending), 'skipped': 0, 'errors': []}
